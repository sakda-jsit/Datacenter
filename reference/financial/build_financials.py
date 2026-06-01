# -*- coding: utf-8 -*-
"""
JSP CONNX - Financial Statement Engine (Prototype / Phase 2 proof)
อ่านงบทดลอง (TB25) -> map บัญชีตามรหัส REF -> ออกงบฐานะการเงิน + งบกำไรขาดทุน
แล้วเทียบกับตัวเลขที่อยู่ในชีต BAL1/BAL2/PL เพื่อพิสูจน์ว่า logic ตรงกับไฟล์เดิม

ผลพลอยได้: generate ไฟล์ seed_statement_mapping.sql (ตาราง account -> REF) ให้ฝั่ง SQL Server
รัน: python build_financials.py
"""
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "..", "2025_JSPC_FIN.xlsx")
SEED_OUT = os.path.join(HERE, "..", "db", "seed_statement_mapping.sql")

# ---- นิยามบรรทัดในงบการเงิน (REF -> ชื่อบรรทัด, หมวด, ลำดับ) -----------------
# section: A=สินทรัพย์ L=หนี้สิน E=ทุน I=รายได้ X=ค่าใช้จ่าย
REF_DEF = {
    # งบฐานะการเงิน - สินทรัพย์
    "A1":  ("เงินสดและรายการเทียบเท่าเงินสด", "A", 11),
    "A2":  ("เงินลงทุนระยะสั้น", "A", 12),
    "A7":  ("ลูกหนี้การค้าและลูกหนี้หมุนเวียนอื่น", "A", 13),
    "A8":  ("ลูกหนี้เงินให้กู้ยืมบุคคล/กิจการที่เกี่ยวข้อง", "A", 14),
    "A3":  ("สินค้าคงเหลือ", "A", 15),
    "A4":  ("สินทรัพย์หมุนเวียนอื่น", "A", 16),
    "A9":  ("เงินลงทุนระยะยาว", "A", 21),
    "A5":  ("ที่ดิน อาคารและอุปกรณ์", "A", 22),
    "A10": ("สินทรัพย์ไม่มีตัวตน", "A", 23),
    "A6":  ("สินทรัพย์ไม่หมุนเวียนอื่น", "A", 24),
    # งบฐานะการเงิน - หนี้สิน
    "L1":  ("เจ้าหนี้การค้าและเจ้าหนี้หมุนเวียนอื่น", "L", 31),
    "L5":  ("ส่วนของหนี้สินตามสัญญาเช่าที่ถึงกำหนดชำระภายในหนึ่งปี", "L", 32),
    "L3":  ("เงินกู้ยืมระยะสั้น", "L", 33),
    "L2":  ("หนี้สินหมุนเวียนอื่น", "L", 34),
    "L6":  ("เงินกู้ยืมระยะยาว", "L", 41),
    "L4":  ("หนี้สินตามสัญญาเช่า", "L", 42),
    # งบฐานะการเงิน - ส่วนของผู้ถือหุ้น
    "C1":  ("ทุนที่ออกและชำระแล้ว", "E", 51),
    "RE":  ("กำไร (ขาดทุน) สะสม", "E", 52),  # = ยอดยกมาบัญชี 32000 + กำไรสุทธิปีปัจจุบัน
    # งบกำไรขาดทุน - รายได้
    "I1":  ("รายได้จากการขาย", "I", 61),
    "I2":  ("รายได้จากการให้บริการ", "I", 62),
    "I3":  ("รายได้ดอกเบี้ย", "I", 63),
    "I4":  ("รายได้อื่น", "I", 64),
    # งบกำไรขาดทุน - ค่าใช้จ่าย
    "C":   ("ต้นทุนขายหรือต้นทุนการให้บริการ", "X", 71),
    "X1":  ("ค่าใช้จ่ายในการขาย", "X", 72),
    "X2":  ("ค่าใช้จ่ายในการบริหาร", "X", 73),
    "X3":  ("ต้นทุนทางการเงิน", "X", 74),
    "X4":  ("ภาษีเงินได้", "X", 75),
}
# บัญชีกำไรสะสมยกมา (เก็บแยกเพื่อคำนวณ RE)
RE_ACCOUNT = "32000"

# ---- ตัวเลขที่งบในไฟล์ระบุไว้ (คอลัมน์ปี 2568) ใช้ verify ----------------------
STATED_2568 = {
    "A1": 907804.59, "A2": 0, "A7": 53763.28, "A8": 0, "A3": 62635.89,
    "A4": 26515.99, "A9": 100079.10, "A5": 409174.15, "A10": 6146.73, "A6": 27185.69,
    "L1": 111870.21, "L5": 11556.73, "L3": 0, "L2": 0, "L6": 1494288.48, "L4": 96394.02,
    "C1": 1000000.0, "RE": -1120804.02,
    "I1": 2396078.94, "I2": 1174585.0, "I3": 1354.98, "I4": 7138.53,
    "C": 2238451.94, "X1": 62914.89, "X2": 709308.53, "X3": -1840.47, "X4": -47470.04,
}


def load_tb(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["TB25"]
    rows = []
    for r in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=15, values_only=True):
        ref, acct, name = r[1], r[2], r[3]
        if acct is None:
            continue
        bf_dr, bf_cr = r[4] or 0, r[5] or 0           # ยอดยกมา
        # Balance หลังปรับปรุง (c14/c15) ถ้าไม่มีใช้ Balance ก่อนปรับปรุง (c9/c10)
        bal_dr = r[13] if r[13] is not None else (r[8] or 0)
        bal_cr = r[14] if r[14] is not None else (r[9] or 0)
        rows.append({
            "ref": (str(ref).strip() if ref else ""),
            "acct": str(acct).strip(),
            "name": (str(name).strip() if name else ""),
            "bf": (bf_dr or 0) - (bf_cr or 0),
            "bal": (bal_dr or 0) - (bal_cr or 0),   # +เดบิต / -เครดิต
        })
    return rows


def build(rows, tax_expense=47470.04):
    # tax_expense: ภาษีเงินได้นิติบุคคล - มาจากโมดูลภาษี (ชีต TAX/ภงด.50) ไม่ใช่จากบัญชีใน TB
    # รวมยอดตาม REF (net = debit - credit)
    line = {}
    re_bf = 0.0
    for x in rows:
        if x["acct"] == RE_ACCOUNT:
            re_bf += x["bf"]            # ยอดยกมาบัญชีกำไรสะสม (ขาดทุนสะสม = ยอดเดบิต)
            continue
        ref = x["ref"]
        if not ref or ref.upper() not in REF_DEF:
            continue
        line[ref.upper()] = line.get(ref.upper(), 0.0) + x["bal"]

    # งบกำไรขาดทุน
    income = {k: -line.get(k, 0.0) for k in ("I1", "I2", "I3", "I4")}   # รายได้ = ยอดเครดิต -> พลิกบวก
    total_income = sum(income.values())
    cogs = line.get("C", 0.0)
    sell = line.get("X1", 0.0)
    admin = line.get("X2", 0.0)
    total_exp = cogs + sell + admin
    profit_before_fin = total_income - total_exp
    fin_cost = line.get("X3", 0.0)             # ดอกเบี้ยจ่าย (เดบิต +)
    profit_before_tax = profit_before_fin - fin_cost
    tax = tax_expense                          # จากโมดูลภาษี (ภงด.50)
    net_profit = profit_before_tax - tax

    # งบฐานะการเงิน
    assets = {k: line.get(k, 0.0) for k in REF_DEF if REF_DEF[k][1] == "A"}
    liab = {k: -line.get(k, 0.0) for k in REF_DEF if REF_DEF[k][1] == "L"}  # หนี้สิน = ยอดเครดิต -> พลิกบวก
    capital = -line.get("C1", 0.0)
    retained = -re_bf + net_profit             # กำไรสะสม = ยกมา(พลิกด้านเป็นทุน) + กำไรสุทธิปีนี้

    computed = dict(assets)
    computed.update(liab)
    computed["C1"] = capital
    computed["RE"] = retained
    # P&L lines: presentation sign ตามที่งบแสดง (X3/X4 แสดงเป็นหักออก = ติดลบ)
    computed["I1"], computed["I2"], computed["I3"], computed["I4"] = (
        income["I1"], income["I2"], income["I3"], income["I4"])
    computed["C"], computed["X1"], computed["X2"] = cogs, sell, admin
    computed["X3"], computed["X4"] = -fin_cost, -tax
    computed["__profit_before_tax__"] = profit_before_tax

    totals = {
        "total_income": total_income,
        "total_exp": total_exp,
        "net_profit": net_profit,
        "total_assets": sum(assets.values()),
        "total_liab": sum(liab.values()),
        "total_equity": capital + retained,
    }
    return computed, totals


def report(computed):
    print("=" * 74)
    print(f"{'REF':5}{'บรรทัดในงบ':<46}{'คำนวณได้':>14} {'งบระบุ':>14} {'ส่วนต่าง':>10}")
    print("=" * 74)
    ok = True
    for ref in sorted(REF_DEF, key=lambda k: REF_DEF[k][2]):
        c = computed.get(ref, 0.0)
        s = STATED_2568.get(ref, 0.0)
        d = c - s
        flag = "" if abs(d) < 0.01 else "  <-- ไม่ตรง"
        if abs(d) >= 0.01:
            ok = False
        print(f"{ref:5}{REF_DEF[ref][0][:44]:<46}{c:>14,.2f} {s:>14,.2f} {d:>10,.2f}{flag}")
    print("=" * 74)
    print("ผลตรวจสอบ:", "ทุกบรรทัดตรงกับไฟล์ Excel เดิม (ผ่าน)" if ok else "พบบรรทัดไม่ตรง")
    return ok


def write_seed(rows):
    seen = set()
    lines = [
        "-- seed_statement_mapping.sql",
        "-- สร้างอัตโนมัติจาก TB25 ในไฟล์ 2025_JSPC_FIN.xlsx (script: build_financials.py)",
        "-- ตารางผูกเลขบัญชี -> รหัสบรรทัดงบการเงิน (REF)",
        "",
        "-- 1) นิยามบรรทัดในงบ",
        "MERGE statement_line AS t USING (VALUES",
    ]
    vals = []
    for ref, (label, sec, order) in sorted(REF_DEF.items(), key=lambda kv: kv[1][2]):
        vals.append(f"  (N'{ref}', N'{label.replace(chr(39), chr(39)*2)}', N'{sec}', {order})")
    lines.append(",\n".join(vals))
    lines += [
        ") AS s(ref_code, line_name, section, sort_order)",
        "  ON t.ref_code = s.ref_code",
        "WHEN MATCHED THEN UPDATE SET line_name=s.line_name, section=s.section, sort_order=s.sort_order",
        "WHEN NOT MATCHED THEN INSERT (ref_code,line_name,section,sort_order) "
        "VALUES (s.ref_code,s.line_name,s.section,s.sort_order);",
        "",
        "-- 2) ผูกเลขบัญชี -> REF",
    ]
    mapping = []
    for x in rows:
        if not x["acct"] or x["acct"] in seen:
            continue
        seen.add(x["acct"])
        ref = x["ref"].upper() if x["ref"] else ""
        if x["acct"] == RE_ACCOUNT:
            ref = "RE"
        if ref not in REF_DEF:
            continue
        nm = x["name"].replace("'", "''")
        mapping.append(f"  (N'{x['acct']}', N'{nm[:60]}', N'{ref}')")
    lines.append("MERGE account_mapping AS t USING (VALUES")
    lines.append(",\n".join(mapping))
    lines += [
        ") AS s(account_code, account_name, ref_code)",
        "  ON t.account_code = s.account_code",
        "WHEN MATCHED THEN UPDATE SET account_name=s.account_name, ref_code=s.ref_code",
        "WHEN NOT MATCHED THEN INSERT (account_code,account_name,ref_code) "
        "VALUES (s.account_code,s.account_name,s.ref_code);",
        "",
    ]
    with open(SEED_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return len(mapping)


if __name__ == "__main__":
    rows = load_tb(XLSX)
    computed, totals = build(rows)
    ok = report(computed)
    print()
    print(f"รวมรายได้     : {totals['total_income']:>16,.2f}  (งบ: 3,579,157.45)")
    print(f"กำไรสุทธิ      : {totals['net_profit']:>16,.2f}  (งบ:   519,171.58)")
    print(f"รวมสินทรัพย์   : {totals['total_assets']:>16,.2f}  (งบ: 1,593,305.42)")
    print(f"รวมหนี้สิน+ทุน : {totals['total_liab'] + totals['total_equity']:>16,.2f}  (งบ: 1,593,305.42)")
    n = write_seed(rows)
    print(f"\nเขียน seed mapping {n} บัญชี -> {os.path.relpath(SEED_OUT, HERE)}")
