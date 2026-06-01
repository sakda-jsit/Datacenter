# -*- coding: utf-8 -*-
"""
รันตัวดึงข้อมูล Express จากโฟลเดอร์ JSIT2016 แล้ว:
  1) สรุปข้อมูลบริษัท + รายการตาราง
  2) ส่งออกผังบัญชี + งบทดลอง (งวดเปิดปัจจุบัน) เป็น CSV
  3) ตรวจสอบความถูกต้อง โดยเทียบยอดคงเหลือกับชีต TB ในไฟล์ Excel
ผลลัพธ์เขียนลง out/ และ log เป็น UTF-8
"""
import os
import csv
import openpyxl
import express_dbf as ex

HERE = os.path.dirname(os.path.abspath(__file__))
FOLDER = r"D:\ExpressI\JSIT2016"
OUT = os.path.join(HERE, "out")
XLSX = os.path.join(HERE, "..", "..", "2025_JSPC_FIN.xlsx")
LOG = os.path.join(OUT, "_run_log.txt")
os.makedirs(OUT, exist_ok=True)
log_fp = open(LOG, "w", encoding="utf-8")


def log(*a):
    s = " ".join(str(x) for x in a)
    print(s, file=log_fp)


def write_csv(path, rows, fields):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


# ---- 1) ข้อมูลบริษัท + ตาราง ----
info = ex.company_info(FOLDER)
log("=== ข้อมูลบริษัท (ISINFO) ===")
for k, v in info.items():
    log(f"  {k}: {v}")

tables = ex.list_tables(FOLDER)
log(f"\n=== ตารางทั้งหมด {len(tables)} ตาราง (ที่มีระเบียน > 0) ===")
for nm, n in tables:
    if isinstance(n, int) and n > 0:
        log(f"  {nm:18} {n:>7} ระเบียน")

# ---- 2) ส่งออกผังบัญชี + งบทดลอง ----
coa = ex.chart_of_accounts(FOLDER)
write_csv(os.path.join(OUT, "chart_of_accounts.csv"), coa,
          ["account_code", "account_name", "account_name2", "level", "parent", "group", "acctype", "status"])
log(f"\nส่งออกผังบัญชี: chart_of_accounts.csv ({len(coa)} บัญชี)")

for ys, label in [("LY", "งวดก่อน"), ("CUR", "งวดเปิดปัจจุบัน"), ("NY", "งวดถัดไป")]:
    tb = ex.trial_balance(FOLDER, ys)
    write_csv(os.path.join(OUT, f"trial_balance_{ys}.csv"), tb,
              ["account_code", "begin_net", "period_debit", "period_credit", "balance_net"])
    log(f"ส่งออกงบทดลอง [{label}]: trial_balance_{ys}.csv ({len(tb)} บัญชี)")

# ---- 3) ตรวจสอบกับ Excel ----
# งวดเปิดปัจจุบันของ JSIT2016 = ปี 2021 -> เทียบกับชีต TB21 (Balance หลังปรับปรุง = c14 - c15)
def excel_tb(sheet):
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb[sheet]
    res = {}
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=15, values_only=True):
        acc = row[2]
        if acc is None:
            continue
        acc = str(acc).strip()
        dr = row[13] if row[13] is not None else (row[8] or 0)
        cr = row[14] if row[14] is not None else (row[9] or 0)
        if acc and (isinstance(dr, (int, float)) or isinstance(cr, (int, float))):
            res[acc] = round((dr or 0) - (cr or 0), 2)
    return res


def verify(year_set, sheet):
    tb = {r["account_code"]: r["balance_net"] for r in ex.trial_balance(FOLDER, year_set)}
    xl = excel_tb(sheet)
    common = sorted(set(tb) & set(xl))
    diffs = [(a, tb[a], xl[a]) for a in common if abs(tb[a] - xl[a]) > 0.01]
    log(f"\n=== ตรวจสอบงบทดลอง [{year_set}] กับชีต {sheet} ===")
    log(f"  บัญชีที่ตรวจ (มีทั้งสองฝั่ง): {len(common)}  | ไม่ตรง: {len(diffs)}")
    for a, t, x in diffs[:15]:
        log(f"    {a}: DBF={t:,.2f}  Excel={x:,.2f}  diff={t-x:,.2f}")
    return len(diffs) == 0, len(common)


ok21, n21 = verify("CUR", "TB21")
log("\nสรุปผลตรวจสอบ:",
    f"งวดปัจจุบัน(2021) ตรงกับ TB21 ทั้งหมด {n21} บัญชี ✓" if ok21
    else "พบบัญชีไม่ตรง (ดูด้านบน)")

log_fp.close()
# echo log to stdout as ascii-safe
print(open(LOG, encoding="utf-8").read())
